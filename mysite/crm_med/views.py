from datetime import timedelta
from django.utils.dateparse import parse_date
from rest_framework import generics, views, status
from .serializers import *
from .models import *
from rest_framework.response import Response
from django.db.models import Q
from rest_framework.views import APIView
from django.utils import timezone
from rest_framework.exceptions import ValidationError
from django.db.models import F, Sum, Value, IntegerField, Case, When, Q
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework.permissions import IsAuthenticated
from django.utils.translation import gettext as _
import openpyxl
from openpyxl.utils import get_column_letter
from django.db.models import Count
from .permissions import *
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status


class CustomLoginView(TokenObtainPairView):
    serializer_class = LoginSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
        except Exception as e:
            return Response({"detail": f"Неверные учетные данные: {e}"}, status=status.HTTP_401_UNAUTHORIZED)

        user = serializer.validated_data
        return Response(serializer.data, status=status.HTTP_200_OK)


class DepartmentPatientAPIView(generics.RetrieveAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentPatientSerializer
    permission_classes = [IsAdmin | IsReceptionist]

    def retrieve(self, request, *args, **kwargs):
        # Получаем департамент
        department = self.get_object()

        # Получаем параметры фильтра
        search_name = request.query_params.get('name')
        doctor_id = request.query_params.get('doctor')
        date_str = request.query_params.get('date')

        # Берём пациентов, связанных с департаментом
        patients_qs = department.patients.select_related('doctor', 'service_type')

        # поиск по имени клиента
        if search_name:
            patients_qs = patients_qs.filter(name=search_name)

        # Фильтрация по доктору
        if doctor_id:
            patients_qs = patients_qs.filter(doctor_id=doctor_id)

        # Фильтрация по дате
        if date_str:
            selected_date = parse_date(date_str)
            if not selected_date:
                raise ValidationError({"date": "Invalid date format, use YYYY-MM-DD"})
            patients_qs = patients_qs.filter(appointment_date__date=selected_date)

        # Временно подменим queryset пациентов на отфильтрованный
        serializer = self.get_serializer(department)
        data = serializer.data

        # Заменим пациентов в ответе на отфильтрованные
        data['patients'] = PatientListSerializer(patients_qs, many=True).data

        return Response(data)


class PatientCreateAPIView(views.APIView):
    permission_classes = [IsAdmin | IsReceptionist]

    def post(self, request):
        serializer = PatientCreateSerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            name = data.get("name")
            patient_name_db = Patient.objects.filter(name__iexact=name).exists()
            serializer.save(primary_patient=not patient_name_db)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PatientEditAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientEditSerializer


class PatientHistoryAPIView(generics.ListAPIView):
    serializer_class = PatientHistoryAppointmentSerializer

    def get_queryset(self):
        patient_name = self.kwargs.get('patient_name')
        qs = Patient.objects.filter(name=patient_name)
        period = self.request.query_params.get('period')
        if not period:
            return qs

        now = timezone.now().date()

        if period == "daily":
            start_date = now - timedelta(days=1)
        elif period == "weekly":
            start_date = now - timedelta(days=7)
        elif period == "monthly":
            start_date = now - timedelta(days=30)
        elif period == "yearly":
            start_date = now - timedelta(days=365)
        else:
            raise ValidationError({"period": "Invalid period. Use daily/weekly/monthly/yearly"})

        qs = qs.filter(created_date__gte=start_date)

        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Считаем статистику по статусам
        status_counts = queryset.values('patient_status').annotate(count=Count('id'))
        counts_dict = {status: 0 for status, _ in PATIENT_STATUS_CHOICES}

        for item in status_counts:
            counts_dict[item['patient_status']] = item['count']

        # Общее количество
        total_count = queryset.count()

        counts_dict['all'] = total_count

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            "report": counts_dict,
            "patients": serializer.data
        })


class PatientHistoryAppointmentAPIView(generics.ListAPIView):
    serializer_class = PatientHistoryAppointmentSerializer

    def get_queryset(self):
        patient_name = self.kwargs.get('patient_name')
        qs = Patient.objects.filter(
            name=patient_name,
            patient_status='had an appointment'
        )

        period = self.request.query_params.get('period')
        if not period:  # если параметр не задан — возвращаем всех
            return qs

        now = timezone.now()

        if period == "daily":
            start_date = now - timedelta(days=1)
        elif period == "weekly":
            start_date = now - timedelta(days=7)
        elif period == "monthly":
            start_date = now - timedelta(days=30)
        elif period == "yearly":
            start_date = now - timedelta(days=365)
        else:
            raise ValidationError({"period": "Invalid period. Use daily/weekly/monthly/yearly"})

        # фильтруем по дате, если period был передан
        return qs.filter(appointment_date__gte=start_date)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        return Response({
            'patient_quantity': queryset.count(),
            'patients': serializer.data,
        })


class PatientHistoryPaymentAPIView(generics.ListAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientHistoryPaymentSerializer

    def get_queryset(self):
        patient_name = self.kwargs.get('patient_name')
        qs = Patient.objects.filter(
            Q(name=patient_name) &
            Q(patient_status='had an appointment')
        )

        period = self.request.query_params.get('period')
        if not period:  # если параметр не задан — возвращаем всех
            return qs

        now = timezone.now()

        if period == "daily":
            start_date = now - timedelta(days=1)
        elif period == "weekly":
            start_date = now - timedelta(days=7)
        elif period == "monthly":
            start_date = now - timedelta(days=30)
        elif period == "yearly":
            start_date = now - timedelta(days=365)
        else:
            raise ValidationError({"period": "Invalid period. Use daily/weekly/monthly/yearly"})

        # фильтруем по дате, если period был передан
        return qs.filter(appointment_date__gte=start_date)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        cash = 0
        card = 0
        for i in queryset:
            price = i.with_discount if i.with_discount else i.service_type.price
            if i.payment_type == 'cash':
                cash += price
            else:
                card += price

        total_sum = cash + card

        return Response({
            'cash': cash,
            'card': card,
            'total_sum': total_sum,
            'patients': serializer.data,
        })


class PatientInfoAPIView(generics.RetrieveAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientInfoSerializer


class ReceptionistEditAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Patient.objects.all()
    serializer_class = ReceptionistSerializer
    permission_classes = [IsAdmin | IsReceptionist]


class DoctorListAPIView(generics.ListAPIView):
    queryset = Doctor.objects.all()
    serializer_class = DoctorListSerializer
    permission_classes = [IsAdmin | IsReceptionist]

    def get_queryset(self):
        qs = Doctor.objects.all()

        # Фильтр по департаменту
        department_id = self.request.query_params.get('department')
        if department_id:
            qs = qs.filter(department=department_id)

        # Фильтр по имени (частичное совпадение, без учёта регистра)
        search_name = self.request.query_params.get('name')
        if search_name:
            qs = qs.filter(username__icontains=search_name)

        return qs


class DoctorEditAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Doctor.objects.all()
    serializer_class = DoctorCreateEditSerializer
    permission_classes = [IsAdmin | DoctorRetrieve]


class DoctorNotificationAPIView(generics.ListAPIView):
    queryset = Patient.objects.all()
    serializer_class = DoctorNotificationSerializer
    permission_classes = [IsDoctor]

    def get_queryset(self):
        return Patient.objects.filter(doctor=self.request.user.id)


class DoctorCreateAPIView(generics.CreateAPIView):
    queryset = Doctor.objects.all()
    serializer_class = DoctorCreateEditSerializer
    permission_classes = [IsAdmin]

    def perform_create(self, serializer):
        serializer.save(user_role="doctor")


class UserProfileAPIView(generics.ListAPIView):
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer


class DepartmentServiceAPIView(generics.ListAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentServicesSerializer
    permission_classes = [IsAdmin | IsReceptionist]


class JobTitleAPIView(generics.ListAPIView):
    queryset = JobTitle.objects.all()
    serializer_class = JobTitleSerializer


class ReportExactAPIView(generics.ListAPIView):
    """
    Возвращает список пациентов, отфильтрованных по:
    - doctor (id доктора, опционально)
    - department (id департамента, опционально)
    - date (YYYY-MM-DD, опционально)

    Если передан параметр export=excel, возвращается Excel-файл.
    """
    serializer_class = ReportExactSerializer
    permission_classes = [IsAdmin | IsReceptionist]

    def get_queryset(self):
        qs = Patient.objects.select_related("service_type", "doctor")

        search_name = self.request.query_params.get('name')
        doctor_id = self.request.query_params.get('doctor')
        department_id = self.request.query_params.get('department')
        date_str = self.request.query_params.get('date')

        if search_name:
            qs = qs.filter(name=search_name)
        if doctor_id:
            qs = qs.filter(doctor_id=doctor_id)
        if department_id:
            qs = qs.filter(department_id=department_id)
        if date_str:
            selected_date = parse_date(date_str)
            if not selected_date:
                raise ValidationError({"date": "Invalid date format, use YYYY-MM-DD"})
            qs = qs.filter(appointment_date__date=selected_date)

        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # --- Если экспорт в Excel ---
        if request.query_params.get("export") == "excel":
            return self.export_to_excel(queryset)

        # --- Подсчёты ---
        sum_discount = queryset.aggregate(
            total=Coalesce(Sum('with_discount', filter=Q(with_discount__isnull=False)), Value(0))
        )['total']

        sum_no_discount = queryset.aggregate(
            total=Coalesce(
                Sum('service_type__price', filter=Q(with_discount__isnull=True)),
                Value(0)
            )
        )['total']

        doctor_earnings = 0
        for patient in queryset:
            price = patient.with_discount if patient.with_discount else patient.service_type.price
            percent = patient.doctor.bonus or 0
            doctor_earnings += price * percent / 100

        patients_count = queryset.count()

        payments = queryset.aggregate(
            total_cash=Coalesce(
                Sum(
                    Case(
                        When(payment_type='cash',
                             then=Case(
                                 When(with_discount__isnull=False, then=F('with_discount')),
                                 default=F('service_type__price'),
                                 output_field=IntegerField()
                             )
                             ),
                        default=Value(0),
                        output_field=IntegerField()
                    )
                ),
                Value(0)
            ),
            total_card=Coalesce(
                Sum(
                    Case(
                        When(payment_type='card',
                             then=Case(
                                 When(with_discount__isnull=False, then=F('with_discount')),
                                 default=F('service_type__price'),
                                 output_field=IntegerField()
                             )
                             ),
                        default=Value(0),
                        output_field=IntegerField()
                    )
                ),
                Value(0)
            )
        )

        serializer = self.get_serializer(queryset, many=True)

        return Response({
            "sum_discount": sum_discount,
            "sum_no_discount": sum_no_discount,
            "doctor_earnings": round(doctor_earnings, 2),
            "patients_count": patients_count,
            "total_cash": payments['total_cash'],
            "total_card": payments['total_card'],
            "patients": serializer.data
        })

    def export_to_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patients"

        # Заголовки по языку
        lang = self.request.LANGUAGE_CODE
        if lang == 'ru':
            headers = [
                "ID", "Дата", "Имя", "Услуга",
                "Тип оплаты", "Цена", "Скидка", "Доктор"
            ]
        else:
            headers = [
                "ID", "Date", "Name", "Service",
                "Payment type", "Price", "Discount", "Doctor"
            ]
        ws.append(headers)

        # Данные
        for p in queryset:
            price = p.with_discount if p.with_discount else p.service_type.price
            ws.append([
                p.id,
                p.appointment_date.strftime('%d-%m-%Y %H:%M'),
                p.name,
                p.service_type.type,
                p.get_payment_type_display(),
                p.service_type.price if p.with_discount is None else "-",
                p.with_discount if p.with_discount else "-",
                p.doctor.username,
            ])

        # Итоговая строка с количеством и суммой
        total_price = sum(
            p.with_discount if p.with_discount else p.service_type.price
            for p in queryset
        )
        ws.append([])
        ws.append([_("Total patients"), queryset.count()])
        ws.append([_("Total amount"), total_price])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = self.request.query_params.get('filename', 'report_exact.xlsx')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


class ReportDoctorAPIView(generics.ListAPIView):
    """
        Возвращает список пациентов для конкретного доктора и даты.
        Фильтры:
        - doctor (id доктора)
        - date (YYYY-MM-DD)
        """
    serializer_class = ReportDoctorSerializer
    permission_classes = [IsAdmin | IsReceptionist]

    def get_queryset(self):
        qs = Patient.objects.all()

        # ! I don't know why we need this search_name
        search_doctor_name = self.request.query_params.get('name')
        doctor_id = self.request.query_params.get('doctor')
        date_str = self.request.query_params.get('date')
        # Tiffany Hall

        if search_doctor_name:
            qs = qs.filter(doctor__username=search_doctor_name)

        if doctor_id:
            qs = qs.filter(doctor_id=doctor_id)

        if date_str:
            selected_date = parse_date(date_str)
            if not selected_date:
                raise ValidationError({"date": "Invalid date format, use YYYY-MM-DD"})
            qs = qs.filter(appointment_date__date=selected_date)

        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Проверяем, нужно ли экспортировать Excel
        if request.query_params.get("export") == "excel":
            return self.export_to_excel(queryset)

        serializer = self.get_serializer(queryset, many=True)

        # Считаем сумму цен
        total_price = sum(
            p.with_discount if p.with_discount else p.service_type.price
            for p in queryset
        )

        return Response({
            "total_price": total_price,
            "results": serializer.data
        })

    def export_to_excel(self, queryset):
        # Создаем книгу и лист
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patients"

        # Заголовки (локализация)
        lang = self.request.LANGUAGE_CODE
        if lang == 'ru':
            headers = ["ID", "Дата", "Имя", "Цена"]
        else:
            headers = ["ID", "Date", "Name", "Price"]

        ws.append(headers)

        # Данные
        for p in queryset:
            price = p.with_discount if p.with_discount else p.service_type.price
            ws.append([
                p.id,
                p.appointment_date.strftime('%d-%m-%Y %H:%M'),
                p.name,
                price,
            ])

        # Итоговая строка с суммой
        total_price = sum(
            p.with_discount if p.with_discount else p.service_type.price
            for p in queryset
        )
        ws.append([])
        ws.append([_("Total"), "", "", total_price])

        # Ответ с Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = self.request.query_params.get('filename', 'report_doctor.xlsx')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


from django.http import HttpResponse
from rest_framework import generics
from rest_framework.exceptions import ValidationError
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Patient


from django.http import HttpResponse
from rest_framework import generics
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Patient


class ReportSummaryAPIView(generics.ListAPIView):
    """
    Выводит итоговые суммы за период,
    а если передан ?export=excel — возвращает Excel файл.
    """
    permission_classes = [IsAdmin | IsReceptionist]

    def get_queryset(self):
        return Patient.objects.none()

    def get_report_data(self):
        qs = Patient.objects.all()

        # ! I don't know why we need this search_name
        search_name = self.request.query_params.get('name')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if search_name:
            qs = qs.filter(name=search_name)

        if date_from:
            selected_date_from = parse_date(date_from)
            if not selected_date_from:
                raise ValidationError({"date_from": "Invalid date format, use YYYY-MM-DD"})
            qs = qs.filter(appointment_date__date__gte=selected_date_from)

        if date_to:
            selected_date_to = parse_date(date_to)
            if not selected_date_to:
                raise ValidationError({"date_to": "Invalid date format, use YYYY-MM-DD"})
            qs = qs.filter(appointment_date__date__lte=selected_date_to)

        doctor_cash = 0
        doctor_card = 0
        clinic_cash = 0
        clinic_card = 0

        for patient in qs:
            price = patient.with_discount if patient.with_discount else patient.service_type.price
            bonus = patient.doctor.bonus

            doctor_part = price / 100 * bonus if bonus else 1

            if patient.payment_type == 'cash':
                doctor_cash += doctor_part
                clinic_cash += price - doctor_part
            else:
                doctor_card += doctor_part
                clinic_card += price - doctor_part

        total_cash = doctor_cash + clinic_cash
        total_card = doctor_card + clinic_card
        total_clinic = clinic_cash + clinic_card
        total_doctor = doctor_cash + doctor_card

        return {
            'doctor_cash': int(doctor_cash),
            'doctor_card': int(doctor_card),

            'clinic_cash': int(clinic_cash),
            'clinic_card': int(clinic_card),

            'total_cash': int(total_cash),
            'total_card': int(total_card),

            'total_clinic': int(total_clinic),
            'total_doctor': int(total_doctor),
        }

    def list(self, request, *args, **kwargs):
        report_data = self.get_report_data()

        # Если экспорт в Excel
        if request.query_params.get('export') == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Report Summary"

            # Заголовки с учетом языка
            lang = self.request.LANGUAGE_CODE
            if lang == 'ru':
                headers = [
                    "Докторам наличными", "Докторам безналичными",
                    "Сумма клинике наличными", "Сумма клинике безналичными",
                    "Общие наличные", "Общие безналичные",
                    "Клинике", "Докторам",
                ]
            else:
                headers = [
                    "Doctor cash", "Doctor non-cash",
                    "Clinic cash", "Clinic non-cash",
                    "Total cash", "Total non-cash",
                    "Total clinic", "Total doctors",
                ]

            ws.append(headers)

            # Данные
            ws.append([
                report_data['doctor_cash'],
                report_data['doctor_card'],
                report_data['clinic_cash'],
                report_data['clinic_card'],
                report_data['total_cash'],
                report_data['total_card'],
                report_data['total_clinic'],
                report_data['total_doctor'],
            ])

            # Автоматическая ширина колонок
            for i, col in enumerate(ws.columns, start=1):
                max_length = 0
                column = get_column_letter(i)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # HTTP-ответ
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = request.query_params.get('filename', 'report_summary.xlsx')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            wb.save(response)
            return response

        # Если Excel не нужен — JSON
        return Response(report_data)


class AnalysisAPIView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        period = request.query_params.get("period", "weekly")

        now = timezone.now()
        # Настраиваем интервалы
        if period == "daily":
            interval = timedelta(hours=2)  # 12 интервалов
            start_date = now - timedelta(days=1)
        elif period == "weekly":
            interval = timedelta(days=1)  # 7 интервалов
            start_date = now - timedelta(days=7)
        elif period == "monthly":
            interval = timedelta(days=2)  # 15 интервалов
            start_date = now - timedelta(days=30)
        elif period == "yearly":
            interval = None  # Для года будем делить на месяцы
            start_date = now - timedelta(days=365)
        else:
            return Response({"error": "Invalid period"}, status=400)

        patients_qs = Patient.objects.filter(appointment_date__gte=start_date, appointment_date__lte=now)

        # кол-во докторов и пациентов.
        total_doctors = Doctor.objects.count()
        total_patients = patients_qs.count()

        # новые и повторные пациенты.
        primary = 0
        for p in patients_qs:
            if p.primary_patient:
                primary += 1

        primary_percent = 0 if not total_patients else primary / total_patients * 100
        repeated_percent = 0 if not total_patients else 100 - primary_percent

        # рост и падение.
        fall = patients_qs.filter(
            patient_status='canceled',
        ).count()
        fall_percent = 0 if not total_patients else fall / total_patients * 100
        rise_percent = 0 if not total_patients else 100 - fall_percent
        print(f'total: {total_patients}, fall: {fall}, f_per: {fall_percent}, r_per: {rise_percent}')

        # Строим интервалы для chart
        chart = []
        if period == "yearly":
            # 12 интервалов по месяцам
            for m in range(1, 13):
                start_month = start_date.replace(month=m, day=1)
                if m == 12:
                    end_month = start_month.replace(year=start_month.year + 1, month=1, day=1)
                else:
                    end_month = start_month.replace(month=m + 1, day=1)
                bucket_qs = patients_qs.filter(
                    appointment_date__gte=start_month,
                    appointment_date__lt=end_month
                )
                had_an_appointment = bucket_qs.exclude(patient_status='canceled').count()
                canceled = bucket_qs.filter(patient_status='canceled').count()
                chart.append({
                    "appointment_date": start_month,
                    "had_an_appointment": had_an_appointment,
                    "canceled": canceled,
                })
        else:
            current_start = start_date
            while current_start < now:
                current_end = current_start + interval
                bucket_qs = patients_qs.filter(
                    appointment_date__gte=current_start,
                    appointment_date__lt=current_end
                )
                had_an_appointment = bucket_qs.exclude(patient_status='canceled').count()
                canceled = bucket_qs.filter(patient_status='canceled').count()
                chart.append({
                    "appointment_date": current_start,
                    "had_an_appointment": had_an_appointment,
                    "canceled": canceled,
                })
                current_start = current_end

        for row in chart:
            row["appointment_date"] = row["appointment_date"].strftime("%d-%m-%Y %H:%M")

        return Response({
            "total_doctors": total_doctors,
            "total_patients": total_patients,
            "new_percent": round(primary_percent),
            "repeated_percent": round(repeated_percent),
            "rise": round(rise_percent),
            "fall": round(fall_percent),
            "chart": chart,
        })


@api_view(['POST'])
def verify_reset_code(request):
    serializer = VerifyResetCodeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Пароль успешно сброшен.'}, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
